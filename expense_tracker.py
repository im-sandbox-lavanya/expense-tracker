#!/usr/bin/env python3
"""
Expense Tracker Application
A simple command-line expense tracking application with CSV export functionality.
"""

import csv
import json
import os
from datetime import datetime
from typing import List, Dict, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class Expense:
    """Represents an expense entry."""
    
    def __init__(self, date: str, category: str, amount: float, description: str):
        self.date = date
        self.category = category
        self.amount = amount
        self.description = description
    
    def to_dict(self) -> Dict:
        """Convert expense to dictionary."""
        return {
            'date': self.date,
            'category': self.category,
            'amount': self.amount,
            'description': self.description
        }
    
    @classmethod
    def from_dict(cls, data: Dict) -> 'Expense':
        """Create expense from dictionary."""
        return cls(
            date=data['date'],
            category=data['category'],
            amount=data['amount'],
            description=data['description']
        )


class ExpenseTracker:
    """Main expense tracker application."""
    
    def __init__(self, data_file: str = 'expenses.json'):
        self.data_file = data_file
        self.expenses: List[Expense] = []
        self.load_expenses()
    
    def load_expenses(self) -> None:
        """Load expenses from JSON file."""
        if os.path.exists(self.data_file):
            try:
                with open(self.data_file, 'r') as f:
                    data = json.load(f)
                    self.expenses = [Expense.from_dict(expense_data) for expense_data in data]
            except (json.JSONDecodeError, KeyError) as e:
                print(f"Error loading expenses: {e}")
                self.expenses = []
    
    def save_expenses(self) -> None:
        """Save expenses to JSON file."""
        try:
            with open(self.data_file, 'w') as f:
                data = [expense.to_dict() for expense in self.expenses]
                json.dump(data, f, indent=2)
        except IOError as e:
            print(f"Error saving expenses: {e}")
    
    def add_expense(self, date: str, category: str, amount: float, description: str) -> None:
        """Add a new expense."""
        expense = Expense(date, category, amount, description)
        self.expenses.append(expense)
        self.save_expenses()
        print(f"Added expense: {amount} for {description}")
    
    def list_expenses(self) -> None:
        """Display all expenses."""
        if not self.expenses:
            print("No expenses recorded.")
            return
        
        print("\nAll Expenses:")
        print("-" * 70)
        print(f"{'Date':<12} {'Category':<15} {'Amount':<10} {'Description'}")
        print("-" * 70)
        
        for expense in self.expenses:
            print(f"{expense.date:<12} {expense.category:<15} {expense.amount:<10.2f} {expense.description}")
        
        total = sum(expense.amount for expense in self.expenses)
        print("-" * 70)
        print(f"Total: ${total:.2f}")
    
    def export_to_csv(self, filename: Optional[str] = None) -> bool:
        """
        Export all expenses to a CSV file.
        
        Args:
            filename: Optional filename for the CSV. If None, uses timestamp.
        
        Returns:
            bool: True if export successful, False otherwise.
        """
        if not self.expenses:
            print("No expenses to export.")
            return False
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"expenses_export_{timestamp}.csv"
        
        try:
            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = ['date', 'category', 'amount', 'description']
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                
                # Write header
                writer.writeheader()
                
                # Write expense data
                for expense in self.expenses:
                    writer.writerow(expense.to_dict())
            
            print(f"Successfully exported {len(self.expenses)} expenses to {filename}")
            return True
            
        except IOError as e:
            print(f"Error exporting to CSV: {e}")
            return False
        except Exception as e:
            print(f"Unexpected error during export: {e}")
            return False
    
    def export_to_excel(self, filename: Optional[str] = None) -> bool:
        """
        Export all expenses to an Excel file.
        
        Args:
            filename: Optional filename for the Excel file. If None, uses timestamp.
        
        Returns:
            bool: True if export successful, False otherwise.
        """
        if not EXCEL_AVAILABLE:
            print("Excel export not available. Please install openpyxl: pip install openpyxl")
            return False
        
        if not self.expenses:
            print("No expenses to export.")
            return False
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"expenses_export_{timestamp}.xlsx"
        
        try:
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Expenses"
            
            # Define headers
            headers = ['Date', 'Category', 'Amount', 'Description']
            
            # Style for headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center")
            
            # Add headers with styling
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Add expense data
            for row, expense in enumerate(self.expenses, 2):
                ws.cell(row=row, column=1, value=expense.date)
                ws.cell(row=row, column=2, value=expense.category)
                ws.cell(row=row, column=3, value=expense.amount)
                ws.cell(row=row, column=4, value=expense.description)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add summary row
            summary_row = len(self.expenses) + 3
            ws.cell(row=summary_row, column=3, value="Total:")
            total_amount = sum(expense.amount for expense in self.expenses)
            total_cell = ws.cell(row=summary_row, column=4, value=total_amount)
            total_cell.font = Font(bold=True)
            
            # Save the workbook
            wb.save(filename)
            
            print(f"Successfully exported {len(self.expenses)} expenses to {filename}")
            return True
            
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return False


def get_user_input(prompt: str, input_type: type = str):
    """Get user input with type conversion and error handling."""
    while True:
        try:
            value = input(prompt)
            if input_type == float:
                return float(value)
            elif input_type == str:
                return value.strip()
            else:
                return input_type(value)
        except ValueError:
            print(f"Invalid input. Please enter a valid {input_type.__name__}.")


def main():
    """Main application loop."""
    tracker = ExpenseTracker()
    
    print("Welcome to Expense Tracker!")
    print("Track your expenses and export them to CSV for analysis.")
    
    while True:
        print("\n" + "="*50)
        print("Expense Tracker Menu:")
        print("1. Add Expense")
        print("2. List All Expenses")
        print("3. Export to CSV")
        print("4. Export to Excel")
        print("5. Exit")
        print("="*50)
        
        choice = get_user_input("Enter your choice (1-5): ")
        
        if choice == '1':
            print("\nAdd New Expense:")
            date = get_user_input("Enter date (YYYY-MM-DD): ")
            category = get_user_input("Enter category: ")
            amount = get_user_input("Enter amount: $", float)
            description = get_user_input("Enter description: ")
            
            tracker.add_expense(date, category, amount, description)
        
        elif choice == '2':
            tracker.list_expenses()
        
        elif choice == '3':
            print("\nExport to CSV:")
            custom_filename = get_user_input("Enter filename (or press Enter for auto-generated): ")
            filename = custom_filename if custom_filename else None
            tracker.export_to_csv(filename)
        
        elif choice == '4':
            print("\nExport to Excel:")
            custom_filename = get_user_input("Enter filename (or press Enter for auto-generated): ")
            filename = custom_filename if custom_filename else None
            tracker.export_to_excel(filename)
        
        elif choice == '5':
            print("Thank you for using Expense Tracker!")
            break
        
        else:
            print("Invalid choice. Please enter 1, 2, 3, 4, or 5.")


if __name__ == "__main__":
    main()