#!/usr/bin/env python3
"""
Test suite for the Expense Tracker application.
"""

import unittest
import os
import csv
import json
import tempfile
from expense_tracker import Expense, ExpenseTracker

# Check if openpyxl is available for Excel tests
try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class TestExpense(unittest.TestCase):
    """Test cases for the Expense class."""
    
    def test_expense_creation(self):
        """Test expense object creation."""
        expense = Expense("2024-01-15", "Food", 25.50, "Lunch at restaurant")
        self.assertEqual(expense.date, "2024-01-15")
        self.assertEqual(expense.category, "Food")
        self.assertEqual(expense.amount, 25.50)
        self.assertEqual(expense.description, "Lunch at restaurant")
    
    def test_expense_to_dict(self):
        """Test expense to dictionary conversion."""
        expense = Expense("2024-01-15", "Food", 25.50, "Lunch at restaurant")
        expected = {
            'date': "2024-01-15",
            'category': "Food",
            'amount': 25.50,
            'description': "Lunch at restaurant"
        }
        self.assertEqual(expense.to_dict(), expected)
    
    def test_expense_from_dict(self):
        """Test expense creation from dictionary."""
        data = {
            'date': "2024-01-15",
            'category': "Food",
            'amount': 25.50,
            'description': "Lunch at restaurant"
        }
        expense = Expense.from_dict(data)
        self.assertEqual(expense.date, "2024-01-15")
        self.assertEqual(expense.category, "Food")
        self.assertEqual(expense.amount, 25.50)
        self.assertEqual(expense.description, "Lunch at restaurant")


class TestExpenseTracker(unittest.TestCase):
    """Test cases for the ExpenseTracker class."""
    
    def setUp(self):
        """Set up test environment."""
        self.temp_dir = tempfile.mkdtemp()
        self.test_data_file = os.path.join(self.temp_dir, 'test_expenses.json')
        self.tracker = ExpenseTracker(self.test_data_file)
    
    def tearDown(self):
        """Clean up test environment."""
        # Clean up any created files
        if os.path.exists(self.test_data_file):
            os.remove(self.test_data_file)
        
        # Clean up CSV files in temp directory
        for file in os.listdir(self.temp_dir):
            if file.endswith('.csv'):
                os.remove(os.path.join(self.temp_dir, file))
    
    def test_add_expense(self):
        """Test adding an expense."""
        self.tracker.add_expense("2024-01-15", "Food", 25.50, "Lunch")
        self.assertEqual(len(self.tracker.expenses), 1)
        expense = self.tracker.expenses[0]
        self.assertEqual(expense.date, "2024-01-15")
        self.assertEqual(expense.category, "Food")
        self.assertEqual(expense.amount, 25.50)
        self.assertEqual(expense.description, "Lunch")
    
    def test_save_and_load_expenses(self):
        """Test saving and loading expenses."""
        # Add expenses
        self.tracker.add_expense("2024-01-15", "Food", 25.50, "Lunch")
        self.tracker.add_expense("2024-01-16", "Transport", 15.00, "Bus fare")
        
        # Create new tracker instance to test loading
        new_tracker = ExpenseTracker(self.test_data_file)
        self.assertEqual(len(new_tracker.expenses), 2)
        
        # Verify data integrity
        expenses = new_tracker.expenses
        self.assertEqual(expenses[0].date, "2024-01-15")
        self.assertEqual(expenses[0].category, "Food")
        self.assertEqual(expenses[1].date, "2024-01-16")
        self.assertEqual(expenses[1].category, "Transport")


class TestCSVExport(unittest.TestCase):
    """Test cases for CSV export functionality."""
    
    def setUp(self):
        """Set up test environment."""
        self.temp_dir = tempfile.mkdtemp()
        self.test_data_file = os.path.join(self.temp_dir, 'test_expenses.json')
        self.tracker = ExpenseTracker(self.test_data_file)
        
        # Add test data
        self.tracker.add_expense("2024-01-15", "Food", 25.50, "Lunch at restaurant")
        self.tracker.add_expense("2024-01-16", "Transport", 15.00, "Bus fare")
        self.tracker.add_expense("2024-01-17", "Entertainment", 45.75, "Movie, popcorn")
        self.tracker.add_expense("2024-01-18", "Food", 12.25, "Coffee, pastry")
    
    def tearDown(self):
        """Clean up test environment."""
        # Clean up files
        for file in os.listdir(self.temp_dir):
            file_path = os.path.join(self.temp_dir, file)
            if os.path.isfile(file_path):
                os.remove(file_path)
    
    def test_export_to_csv_basic(self):
        """Test basic CSV export functionality."""
        csv_file = os.path.join(self.temp_dir, 'test_export.csv')
        result = self.tracker.export_to_csv(csv_file)
        
        self.assertTrue(result)
        self.assertTrue(os.path.exists(csv_file))
        
        # Verify CSV content
        with open(csv_file, 'r', newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            
            self.assertEqual(len(rows), 4)
            
            # Check first row
            self.assertEqual(rows[0]['date'], "2024-01-15")
            self.assertEqual(rows[0]['category'], "Food")
            self.assertEqual(rows[0]['amount'], "25.5")
            self.assertEqual(rows[0]['description'], "Lunch at restaurant")
            
            # Check that all expected columns are present
            expected_columns = {'date', 'category', 'amount', 'description'}
            self.assertEqual(set(rows[0].keys()), expected_columns)
    
    def test_export_to_csv_with_special_characters(self):
        """Test CSV export with special characters in descriptions."""
        # Add expense with special characters
        self.tracker.add_expense("2024-01-19", "Food", 30.00, "Lunch, \"quoted\", with commas")
        
        csv_file = os.path.join(self.temp_dir, 'test_special_chars.csv')
        result = self.tracker.export_to_csv(csv_file)
        
        self.assertTrue(result)
        
        # Verify CSV handles special characters correctly
        with open(csv_file, 'r', newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            
            # Find the row with special characters
            special_row = None
            for row in rows:
                if "quoted" in row['description']:
                    special_row = row
                    break
            
            self.assertIsNotNone(special_row)
            self.assertEqual(special_row['description'], "Lunch, \"quoted\", with commas")
    
    def test_export_empty_expenses(self):
        """Test CSV export with no expenses."""
        empty_tracker = ExpenseTracker(os.path.join(self.temp_dir, 'empty.json'))
        csv_file = os.path.join(self.temp_dir, 'empty_export.csv')
        
        result = empty_tracker.export_to_csv(csv_file)
        self.assertFalse(result)
        self.assertFalse(os.path.exists(csv_file))
    
    def test_export_auto_filename(self):
        """Test CSV export with auto-generated filename."""
        # Change to temp directory to control where file is created
        original_cwd = os.getcwd()
        os.chdir(self.temp_dir)
        
        try:
            result = self.tracker.export_to_csv()
            self.assertTrue(result)
            
            # Check that a CSV file was created
            csv_files = [f for f in os.listdir('.') if f.endswith('.csv')]
            self.assertEqual(len(csv_files), 1)
            
            # Verify the auto-generated file contains correct data
            csv_file = csv_files[0]
            with open(csv_file, 'r', newline='', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                rows = list(reader)
                self.assertEqual(len(rows), 4)
        
        finally:
            os.chdir(original_cwd)
    
    def test_export_large_dataset(self):
        """Test CSV export with a larger dataset."""
        # Add more expenses to test with larger dataset
        categories = ["Food", "Transport", "Entertainment", "Utilities", "Healthcare"]
        
        for i in range(100):
            date = f"2024-01-{(i % 28) + 1:02d}"
            category = categories[i % len(categories)]
            amount = round(10.0 + (i * 1.5), 2)
            description = f"Test expense {i} with description"
            self.tracker.add_expense(date, category, amount, description)
        
        csv_file = os.path.join(self.temp_dir, 'large_export.csv')
        result = self.tracker.export_to_csv(csv_file)
        
        self.assertTrue(result)
        
        # Verify all data is exported
        with open(csv_file, 'r', newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            # Original 4 + 100 new expenses
            self.assertEqual(len(rows), 104)


@unittest.skipIf(not EXCEL_AVAILABLE, "openpyxl not available")
class TestExcelExport(unittest.TestCase):
    """Test cases for Excel export functionality."""
    
    def setUp(self):
        """Set up test environment."""
        self.temp_dir = tempfile.mkdtemp()
        self.test_data_file = os.path.join(self.temp_dir, 'test_expenses.json')
        self.tracker = ExpenseTracker(self.test_data_file)
        
        # Add test data
        self.tracker.add_expense("2024-01-15", "Food", 25.50, "Lunch at restaurant")
        self.tracker.add_expense("2024-01-16", "Transport", 15.00, "Bus fare")
        self.tracker.add_expense("2024-01-17", "Entertainment", 45.75, "Movie, popcorn")
        self.tracker.add_expense("2024-01-18", "Food", 12.25, "Coffee, pastry")
    
    def tearDown(self):
        """Clean up test environment."""
        # Clean up files
        for file in os.listdir(self.temp_dir):
            file_path = os.path.join(self.temp_dir, file)
            if os.path.isfile(file_path):
                os.remove(file_path)
    
    def test_export_to_excel_basic(self):
        """Test basic Excel export functionality."""
        excel_file = os.path.join(self.temp_dir, 'test_export.xlsx')
        result = self.tracker.export_to_excel(excel_file)
        
        self.assertTrue(result)
        self.assertTrue(os.path.exists(excel_file))
        
        # Verify Excel content
        wb = load_workbook(excel_file)
        ws = wb.active
        
        # Check headers
        self.assertEqual(ws.cell(row=1, column=1).value, "Date")
        self.assertEqual(ws.cell(row=1, column=2).value, "Category")
        self.assertEqual(ws.cell(row=1, column=3).value, "Amount")
        self.assertEqual(ws.cell(row=1, column=4).value, "Description")
        
        # Check first row of data
        self.assertEqual(ws.cell(row=2, column=1).value, "2024-01-15")
        self.assertEqual(ws.cell(row=2, column=2).value, "Food")
        self.assertEqual(ws.cell(row=2, column=3).value, 25.50)
        self.assertEqual(ws.cell(row=2, column=4).value, "Lunch at restaurant")
        
        # Check that we have the right number of data rows (4 + 1 header + 1 empty + 1 total = 7 rows used)
        # Data rows: 2, 3, 4, 5 (4 rows)
        # Total row should be at row 7 (4 data rows + 1 header + 2 gap)
        self.assertEqual(ws.cell(row=7, column=3).value, "Total:")
        self.assertEqual(ws.cell(row=7, column=4).value, 98.50)  # 25.50 + 15.00 + 45.75 + 12.25
    
    def test_export_to_excel_with_special_characters(self):
        """Test Excel export with special characters in descriptions."""
        # Add expense with special characters
        self.tracker.add_expense("2024-01-19", "Food", 30.00, "Lunch, \"quoted\", with commas")
        
        excel_file = os.path.join(self.temp_dir, 'test_special_chars.xlsx')
        result = self.tracker.export_to_excel(excel_file)
        
        self.assertTrue(result)
        
        # Verify Excel handles special characters correctly
        wb = load_workbook(excel_file)
        ws = wb.active
        
        # Find the row with special characters (should be row 6, since we added 1 more expense)
        self.assertEqual(ws.cell(row=6, column=4).value, "Lunch, \"quoted\", with commas")
    
    def test_export_empty_expenses_excel(self):
        """Test Excel export with no expenses."""
        empty_tracker = ExpenseTracker(os.path.join(self.temp_dir, 'empty.json'))
        excel_file = os.path.join(self.temp_dir, 'empty_export.xlsx')
        
        result = empty_tracker.export_to_excel(excel_file)
        self.assertFalse(result)
        self.assertFalse(os.path.exists(excel_file))
    
    def test_export_auto_filename_excel(self):
        """Test Excel export with auto-generated filename."""
        # Change to temp directory to control where file is created
        original_cwd = os.getcwd()
        os.chdir(self.temp_dir)
        
        try:
            result = self.tracker.export_to_excel()
            self.assertTrue(result)
            
            # Check that an Excel file was created
            xlsx_files = [f for f in os.listdir('.') if f.endswith('.xlsx')]
            self.assertEqual(len(xlsx_files), 1)
            
            # Verify the auto-generated file contains correct data
            excel_file = xlsx_files[0]
            wb = load_workbook(excel_file)
            ws = wb.active
            
            # Should have 4 data rows + 1 header
            # Check that we have data in the expected rows
            self.assertEqual(ws.cell(row=2, column=1).value, "2024-01-15")
            self.assertEqual(ws.cell(row=5, column=1).value, "2024-01-18")  # Last data row
        
        finally:
            os.chdir(original_cwd)
    
    def test_excel_styling(self):
        """Test that Excel export includes proper styling."""
        excel_file = os.path.join(self.temp_dir, 'test_styling.xlsx')
        result = self.tracker.export_to_excel(excel_file)
        
        self.assertTrue(result)
        
        # Load workbook and check styling
        wb = load_workbook(excel_file)
        ws = wb.active
        
        # Check that header cells have styling
        header_cell = ws.cell(row=1, column=1)
        self.assertTrue(header_cell.font.bold)
        # RGB color includes alpha channel, so check for the color hex value
        self.assertEqual(header_cell.font.color.rgb, "00FFFFFF")
        
        # Check that total row is bold
        total_cell = ws.cell(row=7, column=4)  # Total amount cell
        self.assertTrue(total_cell.font.bold)


if __name__ == '__main__':
    unittest.main()